import random
from openpyxl import Workbook

wb = Workbook()
wb.save('task.xlsx')

name_list = 'Смирнов Иванов Кузнецов Соколов Попов Лебедев Козлов Новиков Морозов Петров Волков Соловьёв Васильев' \
            ' Зайце Павлов Семёнов Голубев Виноградов Богданов Воробьёв Фёдоров Михайлов Беляев Тарасов Белов' \
            ' Комаров Орлов Киселёв Макаров Андреев Ковалёв Ильин Гусев Титов Кузьмин Кудрявцев Баранов Куликов' \
            ' Алексеев Степанов Яковлев Сорокин Сергеев Романов Захаров Борисов Королёв Герасимов Пономарёв' \
            ' Григорьев'.split()


def vol_1_data_filling(sheet_name):
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    act_sheet = wb[sheet_name]
    act_sheet.cell(row=1, column=1).value = 'Фамилия'
    act_sheet.cell(row=1, column=2).value = 'Взнос вовремя'
    act_sheet.cell(row=1, column=3).value = 'Месяц'
    act_sheet.cell(row=1, column=4).value = 'Сумма'

    for i in range(2, len(name_list) + 2):
        act_sheet.cell(row=i, column=1).value = name_list[i-2]
        act_sheet.cell(row=i, column=2).value = random.choice(['+', '-'])
        act_sheet.cell(row=i, column=3).value = random.choice(['сентябрь', 'октябрь', 'ноябрь'])
        act_sheet.cell(row=i, column=4).value = random.choice([20, 40])


def vol_2_data_filling(sheet_name, data_sheet):
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    act_sheet = wb[sheet_name]
    data_sheet = wb[data_sheet]
    act_sheet.cell(row=1, column=1).value = 'Фамилия'
    act_sheet.cell(row=1, column=2).value = 'Турбаза'
    act_sheet.cell(row=1, column=3).value = 'Статус'

    for i in range(2, data_sheet.max_row):
        act_sheet.cell(row=i, column=1).value = data_sheet.cell(row=i, column=1).value

        if data_sheet.cell(row=i, column=3).value == 'сентябрь':
            if data_sheet.cell(row=i, column=4).value == '20':
                act_sheet.cell(row=i, column=2).value = 'Кит'
            else:
                act_sheet.cell(row=i, column=2).value = 'Кит+'
        elif data_sheet.cell(row=i, column=3).value == 'октябрь':
            if data_sheet.cell(row=i, column=4).value == '20':
                act_sheet.cell(row=i, column=2).value = 'Озон'
            else:
                act_sheet.cell(row=i, column=2).value = 'Озон+'
        else:
            if data_sheet.cell(row=i, column=4).value == '20':
                act_sheet.cell(row=i, column=2).value = 'Бухта'
            else:
                act_sheet.cell(row=i, column=2).value = 'Бухта+'

        if data_sheet.cell(row=i, column=2).value == '+':
            act_sheet.cell(row=i, column=3).value = 'Нет задолженностей'
        else:
            act_sheet.cell(row=i, column=3).value = 'Должник'


vol_1_data_filling('first_list')
vol_2_data_filling('second_list', 'first_list')

wb.save('task.xlsx')
